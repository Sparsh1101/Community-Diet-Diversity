from enum import unique
import io
import os
from django.http.response import FileResponse
import openpyxl
import xlsxwriter
from datetime import datetime
from django.shortcuts import render, redirect
from django.http import HttpResponse, Http404
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required, user_passes_test
from .decorators import *
from .models import *
from .forms import *
from .helper_functions import *
from shared.encryption import EncryptionHelper
from django.utils import timezone
from django.conf import settings

encryptionHelper = EncryptionHelper()


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def teacherAllSessions(request):
    teacher = TeacherInCharge.objects.filter(user=request.user).first()
    objects = Teacher_Session.objects.filter(teacher=teacher)
    sessions = []
    for object in objects:
        sessions.append(object.session)
    upcoming_sessions = []
    open_sessions = []
    close_sessions = []
    open = False
    close = False
    upcoming = False
    for session in sessions:
        if session.start_date > timezone.now():
            upcoming_sessions.append(session)
            upcoming = True
        elif session.end_date == None:
            open_sessions.append(session)
            open = True
        else:
            close_sessions.append(session)
            close = True
    return render(
        request,
        "teacher/teacher_all_sessions.html",
        {
            "open_sessions": open_sessions,
            "upcoming_sessions": upcoming_sessions,
            "close_sessions": close_sessions,
            "open": open,
            "upcoming": upcoming,
            "close": close,
            "page_type": "teacher_all_sessions",
        },
    )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def viewSessionStudents(request, id, open_id, my_messages=None):
    teacher = TeacherInCharge.objects.filter(user=request.user).first()
    session = Session.objects.filter(id=id).first()
    objects = Student_Session.objects.filter(session=session, teacher=teacher)
    students = []
    for object in objects:
        students.append(object.student)
    for student in students:
        student.fname = encryptionHelper.decrypt(student.fname)
        student.lname = encryptionHelper.decrypt(student.lname)
        student.unique_no = encryptionHelper.decrypt(student.unique_no)
    if my_messages != None:
        return render(
            request,
            "teacher/view_session_students.html",
            {
                "students": students,
                "session": session,
                "my_messages": my_messages,
                "page_type": "view_session_students",
                "open_id": open_id,
            },
        )
    else:
        return render(
            request,
            "teacher/view_session_students.html",
            {
                "students": students,
                "session": session,
                "page_type": "view_session_students",
                "open_id": open_id,
            },
        )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def getSessionStudentsTemplate(request):
    output = io.BytesIO()
    wb = xlsxwriter.Workbook(output)
    ws = wb.add_worksheet("Session Students Data")
    columns = [
        "Student Username",
    ]
    for col_num in range(len(columns)):
        ws.write(0, col_num, columns[col_num])
    wb.close()
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=sessionStudentTemplate.xlsx"
    return response


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def addSessionStudents(request, id):
    if request.method == "GET":
        return render(
            request,
            "teacher/add_session_students.html",
            {"page_type": "add_session_students"},
        )
    else:
        try:
            excel_file = request.FILES["excel_file"]
            if excel_file.name[-4:] == "xlsx":
                try:
                    wb = openpyxl.load_workbook(excel_file)
                    studentSheet = wb["Session Students Data"]
                except:
                    return render(
                        request,
                        "teacher/add_session_students.html",
                        {
                            "page_type": "add_session_students",
                            "my_messages": {
                                "error": "Incorrect file uploaded, please use the template provided above."
                            },
                        },
                    )
            else:
                return render(
                    request,
                    "teacher/add_session_students.html",
                    {
                        "page_type": "add_session_students",
                        "my_messages": {
                            "error": "Incorrect file type, only .xlsx files are allowed!"
                        },
                    },
                )
        except:
            return render(
                request,
                "teacher/add_session_students.html",
                {
                    "page_type": "add_session_students",
                    "my_messages": {"error": "Sorry something went wrong!"},
                },
            )

        organization = (
            TeacherInCharge.objects.filter(user=request.user).first().organization
        )
        breaking = error = False
        error_message = ""
        student_data = []
        for row_no, row in enumerate(studentSheet.iter_rows()):
            if breaking == True:
                break
            if row_no == 0:
                continue
            for cell in row:
                if cell.column_letter == "A":
                    if cell.value == None:
                        breaking = True
                        break
                    else:
                        if User.objects.filter(username=cell.value).exists():
                            user = User.objects.filter(username=cell.value).first()
                            if is_student(user):
                                student_user = StudentsInfo.objects.filter(
                                    user=user
                                ).first()
                                if student_user.organization == organization:
                                    student_data.append(student_user)
                                else:
                                    breaking = error = True
                                    error_message = (
                                        "Student at row number "
                                        + str(row_no + 1)
                                        + " does not belong to your organization"
                                    )
                                    break
                            else:
                                breaking = error = True
                                error_message = (
                                    "User at row number "
                                    + str(row_no + 1)
                                    + " is not a student"
                                )
                                break
                        else:
                            breaking = error = True
                            error_message = "Invalid username at row number " + str(
                                row_no + 1
                            )
                            break

        if error == True:
            return render(
                request,
                "teacher/add_session_students.html",
                {
                    "page_type": "add_session_students",
                    "my_messages": {"error": error_message},
                },
            )

        session = Session.objects.filter(id=id).first()
        student_added = 0
        student_exists = 0
        for student_user in student_data:
            if student_user.session != None:
                student_exists += 1
            else:
                teacher = TeacherInCharge.objects.filter(user=request.user).first()
                student_user.session = session
                student_user.teacher = TeacherInCharge.objects.filter(
                    user=request.user
                ).first()
                student_session = Student_Session()
                student_session.session = session
                student_session.student = student_user
                student_session.teacher = teacher
                student_user.save()
                student_session.save()
                student_added += 1

        if student_added == 0:
            my_messages = {
                "success": "Registration successful. Already registered: "
                + str(student_exists)
            }
        elif student_exists == 0:
            my_messages = {
                "success": "Registration successful. Newly registered: "
                + str(student_added)
            }
        else:
            my_messages = {
                "success": "Registration successful. Already registered: "
                + str(student_exists)
                + ", Newly registered: "
                + str(student_added)
            }

        return viewSessionStudents(request, id, 1, my_messages)


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def addSessionStudentsList(request, id):
    if request.method == "GET":
        students = StudentsInfo.objects.filter(session=None, teacher=None)
        for student in students:
            student.fname = encryptionHelper.decrypt(student.fname)
            student.lname = encryptionHelper.decrypt(student.lname)
            student.unique_no = encryptionHelper.decrypt(student.unique_no)
        return render(
            request,
            "teacher/add_session_students_list.html",
            {
                "page_type": "add_session_students_list",
                "students": students,
            },
        )
    else:
        students_id_list = request.POST.getlist("chk")
        session = Session.objects.filter(id=id).first()
        teacher = TeacherInCharge.objects.filter(user=request.user).first()
        for s_id in students_id_list:
            student = StudentsInfo.objects.filter(id=s_id).first()
            student.session = session
            student.teacher = teacher
            student_session = Student_Session()
            student_session.session = session
            student_session.student = student
            student_session.teacher = teacher
            student_session.save()
            student.save()

        return redirect("accounts:view_session_students", id, 1)


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def removeSessionStudent(request, session_id, student_id):
    student = StudentsInfo.objects.filter(id=student_id).first()
    student.session = None
    student.teacher = None
    Student_Session.objects.filter(student=student).delete()
    student.save()
    my_messages = {"success": "Student removed from session successfully"}
    return viewSessionStudents(request, session_id, 1, my_messages)


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def viewSessionForms(request, id):
    teacher = TeacherInCharge.objects.get(user=request.user)
    teachersession = Session.objects.filter(id=id).first()
    total_students = StudentsInfo.objects.filter(
        session=teachersession, teacher=teacher
    )
    results = []
    closed_sessions = FormDetails.objects.filter(
        teacher=teacher, session=teachersession, open=False
    )
    closed_info_sessions = InfoFormDetails.objects.filter(open=False)
    for session in closed_sessions:
        temp_list = [session.form, session.start_timestamp, session.end_timestamp]
        if session.pre:
            temp_list.append("Pre Test")
        else:
            temp_list.append("Post Test")
        count = 0
        for student in total_students:
            if ModuleOne.objects.filter(
                student=student,
                submission_timestamp__gte=session.start_timestamp,
                submission_timestamp__lte=session.end_timestamp,
            ).exists():
                draftForm = ModuleOne.objects.filter(
                    student=student,
                    submission_timestamp__gte=session.start_timestamp,
                    submission_timestamp__lte=session.end_timestamp,
                ).first()
                if not draftForm.draft:
                    count += 1

        temp_list.append(count)
        temp_list.append(len(total_students))
        temp_list.append(session.id)
        results.append(temp_list)

    for session in closed_info_sessions:
        temp_list = [session.form, session.start_timestamp, session.end_timestamp]
        temp_list.append("None")
        count = 0
        for student in total_students:
            if Physique.objects.filter(
                student=student,
                submission_timestamp__gte=session.start_timestamp,
                submission_timestamp__lte=session.end_timestamp,
            ).exists():
                draftForm = Physique.objects.filter(
                    student=student,
                    submission_timestamp__gte=session.start_timestamp,
                    submission_timestamp__lte=session.end_timestamp,
                ).first()
                if not draftForm.draft:
                    count += 1

        temp_list.append(count)
        temp_list.append(len(total_students))
        temp_list.append(session.id)
        results.append(temp_list)

    open_sessions = FormDetails.objects.filter(
        teacher=teacher, session=teachersession, open=True
    )
    open_info_sessions = InfoFormDetails.objects.filter(open=True)
    results2 = []
    for session in open_sessions:
        temp_list = [session.form, session.start_timestamp]
        if session.pre:
            temp_list.append("Pre Test")
        else:
            temp_list.append("Post Test")
        count = 0
        for student in total_students:
            if ModuleOne.objects.filter(
                student=student, submission_timestamp__gte=session.start_timestamp
            ).exists():
                draftForm = ModuleOne.objects.filter(
                    student=student, submission_timestamp__gte=session.start_timestamp
                ).first()
                if not draftForm.draft:
                    count += 1

        temp_list.append(count)
        temp_list.append(len(total_students))
        temp_list.append(session.id)
        results2.append(temp_list)

    for session in open_info_sessions:
        temp_list = [session.form, session.start_timestamp]
        temp_list.append("None")
        count = 0
        for student in total_students:
            if Physique.objects.filter(
                student=student, submission_timestamp__gte=session.start_timestamp
            ).exists():
                draftForm = Physique.objects.filter(
                    student=student, submission_timestamp__gte=session.start_timestamp
                ).first()
                if not draftForm.draft:
                    count += 1

        temp_list.append(count)
        temp_list.append(len(total_students))
        temp_list.append(session.id)
        results2.append(temp_list)
    return render(
        request,
        "teacher/view_session_forms.html",
        {
            "results": results,
            "results2": results2,
            "page_type": "view_session_forms",
            "session": teachersession,
        },
    )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def getFormDetails(request, id, session_id, form_type):
    if form_type == 1:
        form = FormDetails.objects.get(pk=id)
    else:
        form = InfoFormDetails.objects.get(pk=id)
    teacher = TeacherInCharge.objects.get(user=request.user)
    session = Session.objects.filter(id=session_id).first()
    session_students = Student_Session.objects.filter(teacher=teacher, session=session)
    total_students = []
    for session_student in session_students:
        total_students.append(session_student.student)

    filled_students = []
    not_filled_students = []
    if not form.open:
        form_open = False
        temp_list = [form.form, form.start_timestamp, form.end_timestamp]
    else:
        form_open = True
        temp_list = [form.form, form.start_timestamp]

    if form_type == 1:
        if form.pre:
            temp_list.append("Pre Test")
        else:
            temp_list.append("Post Test")
    else:
        temp_list.append("None")

    if form.form.name == "moduleOne":
        for student in total_students:
            temp = []
            if form.open:
                if ModuleOne.objects.filter(
                    student=student, submission_timestamp__gte=form.start_timestamp
                ).exists():
                    draftForm = ModuleOne.objects.filter(
                        student=student, submission_timestamp__gte=form.start_timestamp
                    ).first()

                    if draftForm.draft:
                        temp.append(
                            encryptionHelper.decrypt(student.fname)
                            + " "
                            + encryptionHelper.decrypt(student.lname)
                        )
                        temp.append("-")
                        not_filled_students.append(temp)
                    else:
                        temp.append(
                            encryptionHelper.decrypt(student.fname)
                            + " "
                            + encryptionHelper.decrypt(student.lname)
                        )
                        temp.append(draftForm.submission_timestamp)
                        filled_students.append(temp)
                else:
                    temp.append(
                        encryptionHelper.decrypt(student.fname)
                        + " "
                        + encryptionHelper.decrypt(student.lname)
                    )
                    temp.append("-")
                    not_filled_students.append(temp)

            else:
                if ModuleOne.objects.filter(
                    student=student,
                    submission_timestamp__gte=form.start_timestamp,
                    submission_timestamp__lte=form.end_timestamp,
                ).exists():
                    submitted_form = ModuleOne.objects.filter(
                        student=student, submission_timestamp__gte=form.start_timestamp
                    ).first()
                    temp.append(
                        encryptionHelper.decrypt(student.fname)
                        + " "
                        + encryptionHelper.decrypt(student.lname)
                    )
                    temp.append(submitted_form.submission_timestamp)
                    filled_students.append(temp)
                else:
                    temp.append(
                        encryptionHelper.decrypt(student.fname)
                        + " "
                        + encryptionHelper.decrypt(student.lname)
                    )
                    temp.append("-")
                    not_filled_students.append(temp)

    elif form.form.name == "physique":
        for student in total_students:
            temp = []
            if Physique.objects.filter(
                student=student, submission_timestamp__gte=form.start_timestamp
            ).exists():
                draftForm = Physique.objects.filter(
                    student=student, submission_timestamp__gte=form.start_timestamp
                ).first()

                if draftForm.draft:
                    temp.append(
                        encryptionHelper.decrypt(student.fname)
                        + " "
                        + encryptionHelper.decrypt(student.lname)
                    )
                    temp.append("-")
                    not_filled_students.append(temp)
                else:
                    temp.append(
                        encryptionHelper.decrypt(student.fname)
                        + " "
                        + encryptionHelper.decrypt(student.lname)
                    )
                    temp.append(draftForm.submission_timestamp)
                    filled_students.append(temp)
            else:
                temp.append(
                    encryptionHelper.decrypt(student.fname)
                    + " "
                    + encryptionHelper.decrypt(student.lname)
                )
                temp.append("-")
                not_filled_students.append(temp)

    return render(
        request,
        "teacher/teacher_dashboard_getDetails.html",
        {
            "result": temp_list,
            "filled_students": filled_students,
            "not_filled_students": not_filled_students,
            "open": form_open,
        },
    )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def manageForms(request, id):
    if request.method == "GET":
        physique = {}
        moduleOne = {}
        moduleTwo = {}
        moduleThree = {}
        teacher = TeacherInCharge.objects.get(user=request.user)
        session = Session.objects.filter(id=id).first()
        form = Form.objects.get(name="moduleOne")
        if FormDetails.objects.filter(
            form=form, teacher=teacher, session=session
        ).exists():
            form = (
                FormDetails.objects.filter(
                    form=form, teacher=teacher, session=session, open=True
                )
                .order_by("-start_timestamp")
                .first()
            )
            if form:
                if form.pre:
                    moduleOne["pre"] = True
                    moduleOne["post"] = False

                else:
                    moduleOne["post"] = True
                    moduleOne["pre"] = False

        physique_form = Form.objects.get(name="physique")
        if FormDetails.objects.filter(
            form=physique_form, teacher=teacher, session=session
        ).exists():
            form2 = (
                FormDetails.objects.filter(
                    form=physique_form, teacher=teacher, session=session, open=True
                )
                .order_by("-start_timestamp")
                .first()
            )

        return render(
            request,
            "teacher/manage_forms_teacher.html",
            {
                "physique": physique,
                "moduleOne": moduleOne,
                "moduleTwo": moduleTwo,
                "moduleThree": moduleThree,
                "page_type": "manage_forms",
            },
        )
    else:

        if "moduleOne" in request.POST:
            module_one_pre = request.POST.get("module_one_pre", False)
            module_one_post = request.POST.get("module_one_post", False)
            if module_one_pre == "on" and module_one_post == "on":
                messages.error(request, "Cannot select both PreTest and PostTest")
                return redirect("accounts:manage_forms")

            form = Form.objects.get(name="moduleOne")
            teacher = TeacherInCharge.objects.get(user=request.user)
            session = Session.objects.filter(id=id).first()
            if module_one_pre == "on":
                if not FormDetails.objects.filter(
                    form=form, teacher=teacher, session=session, pre=True, open=True
                ).exists():
                    update = FormDetails(
                        form=form,
                        teacher=teacher,
                        session=session,
                        pre=True,
                        open=True,
                        start_timestamp=datetime.now(),
                    )
                    update.save()
            else:
                if FormDetails.objects.filter(
                    form=form, teacher=teacher, session=session, pre=True, open=True
                ).exists():
                    update = FormDetails.objects.filter(
                        form=form, teacher=teacher, session=session, pre=True, open=True
                    ).first()
                    update.open = False
                    update.end_timestamp = datetime.now()
                    session = Session.objects.filter(id=id).first()
                    teacher = TeacherInCharge.objects.filter(user=request.user).first()
                    total_students = StudentsInfo.objects.filter(
                        session=session, teacher=teacher
                    )
                    for student in total_students:
                        if ModuleOne.objects.filter(
                            student=student,
                            submission_timestamp__gte=update.start_timestamp,
                            submission_timestamp__lte=update.end_timestamp,
                            draft=True,
                            pre=True,
                        ).exists():
                            draftForm = ModuleOne.objects.filter(
                                student=student,
                                submission_timestamp__gte=update.start_timestamp,
                                submission_timestamp__lte=update.end_timestamp,
                                draft=True,
                                pre=True,
                            ).first()
                            draftForm.delete()
                    update.save()

            if module_one_post == "on":
                if not FormDetails.objects.filter(
                    form=form, teacher=teacher, session=session, pre=False, open=True
                ).exists():
                    update = FormDetails(
                        form=form,
                        teacher=teacher,
                        session=session,
                        pre=False,
                        open=True,
                        start_timestamp=datetime.now(),
                    )
                    update.save()
            else:
                if FormDetails.objects.filter(
                    form=form, teacher=teacher, session=session, pre=False, open=True
                ).exists():
                    update = FormDetails.objects.filter(
                        form=form,
                        teacher=teacher,
                        session=session,
                        pre=False,
                        open=True,
                    ).first()
                    update.open = False
                    update.end_timestamp = datetime.now()
                    teacher = TeacherInCharge.objects.get(user=request.user)
                    session = Session.objects.filter(id=id).first()
                    total_students = StudentsInfo.objects.filter(
                        session=session, teacher=teacher
                    )
                    for student in total_students:
                        if ModuleOne.objects.filter(
                            student=student,
                            submission_timestamp__gte=update.start_timestamp,
                            submission_timestamp__lte=update.end_timestamp,
                            draft=True,
                            pre=False,
                        ).exists():
                            draftForm = ModuleOne.objects.filter(
                                student=student,
                                submission_timestamp__gte=update.start_timestamp,
                                submission_timestamp__lte=update.end_timestamp,
                                draft=True,
                                pre=False,
                            ).first()
                            draftForm.delete()
                    update.save()

        elif "physique" in request.POST:
            physique_pre = request.POST.get("physique_pre", False)
            form = Form.objects.get(name="physique")
            teacher = TeacherInCharge.objects.get(user=request.user)
            session = Session.objects.filter(id=id).first()
            total_students = StudentsInfo.objects.filter(
                session=session, teacher=teacher
            )

            # if physique_pre == "on":
            if not FormDetails.objects.filter(
                form=form, teacher=teacher, session=session, pre=True, open=True
            ).exists():
                update = FormDetails(
                    form=form,
                    teacher=teacher,
                    session=session,
                    pre=True,
                    open=True,
                    start_timestamp=datetime.now(),
                )
                update.save()
            else:
                if FormDetails.objects.filter(
                    form=form, teacher=teacher, session=session, pre=True, open=True
                ).exists():
                    update = FormDetails.objects.filter(
                        form=form, teacher=teacher, session=session, pre=True, open=True
                    ).first()
                    update.open = False
                    update.end_timestamp = datetime.now()
                    teacher = TeacherInCharge.objects.get(user=request.user)
                    session = Session.objects.filter(id=id).first()
                    total_students = StudentsInfo.objects.filter(
                        session=session, teacher=teacher
                    )
                    for student in total_students:
                        if Physique.objects.filter(
                            student=student,
                            submission_timestamp__gte=update.start_timestamp,
                            submission_timestamp__lte=update.end_timestamp,
                            draft=True,
                            pre=True,
                        ).exists():
                            draftForm = Physique.objects.filter(
                                student=student,
                                submission_timestamp__gte=update.start_timestamp,
                                submission_timestamp__lte=update.end_timestamp,
                                draft=True,
                                pre=True,
                            ).first()
                            draftForm.delete()
                    update.save()

            # if physique_post == "on":
            #     if not FormDetails.objects.filter(
            #         form=form, teacher=teacher, session=session, pre=False, open=True
            #     ).exists():
            #         update = FormDetails(
            #             form=form,
            #             teacher=teacher,
            #             session=session,
            #             pre=False,
            #             open=True,
            #             start_timestamp=datetime.now(),
            #         )
            #         update.save()
            # else:
            #     if FormDetails.objects.filter(
            #         form=form, teacher=teacher, session=session, pre=False, open=True
            #     ).exists():
            #         update = FormDetails.objects.filter(
            #             form=form,
            #             teacher=teacher,
            #             session=session,
            #             pre=False,
            #             open=True,
            #         ).first()
            #         update.open = False
            #         update.end_timestamp = datetime.now()
            #         teacher = TeacherInCharge.objects.get(user=request.user)
            #         session = Session.objects.filter(id=id).first()
            #         total_students = StudentsInfo.objects.filter(
            #             session=session, teacher=teacher
            #         )
            #         for student in total_students:
            #             if Physique.objects.filter(
            #                 student=student,
            #                 submission_timestamp__gte=update.start_timestamp,
            #                 submission_timestamp__lte=update.end_timestamp,
            #                 draft=True,
            #                 pre=False,
            #             ).exists():
            #                 draftForm = Physique.objects.filter(
            #                     student=student,
            #                     submission_timestamp__gte=update.start_timestamp,
            #                     submission_timestamp__lte=update.end_timestamp,
            #                     draft=True,
            #                     pre=False,
            #                 ).first()
            #                 draftForm.delete()
            #         update.save()

        module_two_pre = request.POST.get("module_two_pre", False)
        module_two_post = request.POST.get("module_two_post", False)
        module_three_pre = request.POST.get("module_three_pre", False)
        module_three_post = request.POST.get("module_three_post", False)

        return redirect("accounts:manage_forms", id)


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def getTemplate(request):
    file_path = os.path.join(
        settings.BASE_DIR, "accounts/Bulk Registration Template.xlsx"
    )
    if os.path.exists(file_path):
        with open(file_path, "rb") as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
            response["Content-Disposition"] = "inline; filename=" + os.path.basename(
                file_path
            )
            return response
    raise Http404


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_teacher, login_url="accounts:forbidden")
@password_change_required
def bulkRegister(request):
    if request.method == "GET":
        return render(
            request,
            "teacher/bulkregistration.html",
            {"page_type": "bulk_register"},
        )
    else:
        try:
            excel_file = request.FILES["excel_file"]
            if excel_file.name[-4:] == "xlsx":
                wb = openpyxl.load_workbook(excel_file)
                parentSheet = wb["Parents Data"]
                studentSheet = wb["Students Data"]
            else:
                return render(
                    request,
                    "teacher/bulkregistration.html",
                    {
                        "page_type": "bulk_register",
                        "my_messages": {
                            "error": "Incorrect file type, only .xlsx files are allowed!"
                        },
                    },
                )
        except:
            return render(
                request,
                "teacher/bulkregistration.html",
                {
                    "page_type": "bulk_register",
                    "my_messages": {
                        "error": "Sorry something went wrong! Please check if the file uploaded is the correct file with proper inputs."
                    },
                },
            )

        breaking = error = False
        error_message = ""
        parent_data = []
        for row_no, row in enumerate(parentSheet.iter_rows()):
            if breaking == True:
                break
            if row_no == 0 or row_no == 1:
                continue
            row_data = []
            for cell in row:
                if cell.column_letter == "A":
                    if cell.value == None:
                        breaking = True
                        break
                    elif cell.value != row_no - 1:
                        breaking = error = True
                        error_message = (
                            "There something wrong in the 'Ref ID' column in 'Parents Data' sheet at row number "
                            + str(row_no + 1)
                        )
                        break
                elif cell.column_letter == "B":
                    if cell.value == "-" or cell.value == None:
                        row_data.append(None)
                    else:
                        if User.objects.filter(username=str(cell.value)).exists():
                            row_data.append(str(cell.value))
                        else:
                            breaking = error = True
                            error_message = (
                                "Parent user with the given 'Username' at row number "
                                + str(row_no + 1)
                                + " does not exist"
                            )
                            break
                elif cell.column_letter == "C":
                    if cell.value == None:
                        breaking = True
                        break
                    if not valid_name(cell.value):
                        breaking = error = True
                        error_message = (
                            "Invalid parent's 'First Name' at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        row_data.append(str(cell.value))
                elif cell.column_letter == "D":
                    if cell.value == None:
                        row_data.append(None)
                    elif not valid_name(cell.value):
                        breaking = error = True
                        error_message = (
                            "Invalid parent's 'Middle Name' at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        row_data.append(str(cell.value))
                elif cell.column_letter == "E":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Parent's 'Last Name' missing at row number "
                            + str(row_no + 1)
                        )
                        break
                    elif not valid_name(cell.value):
                        breaking = error = True
                        error_message = (
                            "Invalid parent's 'Last Name' at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        row_data.append(str(cell.value))
                elif cell.column_letter == "F":
                    if cell.value == None:
                        row_data.append(None)
                    elif not valid_aadhar(str(cell.value)):
                        breaking = error = True
                        error_message = (
                            "Invalid parent's 'Aadhar number' at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        row_data.append(
                            str(cell.value)[:4]
                            + " "
                            + str(cell.value)[4:8]
                            + " "
                            + str(cell.value)[8:]
                        )
                elif cell.column_letter == "G":
                    if cell.value == None:
                        row_data.append(None)
                    elif not valid_email(cell.value):
                        breaking = error = True
                        error_message = "Invalid parent's 'Email' at row number " + str(
                            row_no + 1
                        )
                        break
                    else:
                        row_data.append(str(cell.value))
                elif cell.column_letter == "H":
                    if cell.value == None:
                        row_data.append(None)
                    elif not valid_mobile_no(cell.value):
                        breaking = error = True
                        error_message = (
                            "Invalid parent's 'Mobile Number' at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        row_data.append(str(cell.value))
                elif cell.column_letter == "I":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Parent's 'Gender' missing at row number " + str(row_no + 1)
                        )
                        break
                    else:
                        if check_gender(str(cell.value).capitalize()):
                            row_data.append(str(cell.value).capitalize())
                        else:
                            breaking = error = True
                            error_message = (
                                "Invalid 'Gender' input for 'Parents Data' at row number  "
                                + str(row_no + 1)
                                + ", please select an option from the dropdown menu"
                            )
                            break
                elif cell.column_letter == "J":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Parent's 'Date of Birth' missing at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        dob = (
                            str(cell.value)[8:10]
                            + "/"
                            + str(cell.value)[5:7]
                            + "/"
                            + str(cell.value)[:4]
                        )
                        if (
                            re.match("^[0-9]{2}/[0-9]{2}/[0-9]{4}$", dob)
                            or re.match("^[0-9]{2}-[0-9]{2}-[0-9]{4}$", dob)
                        ) and valid_date(dob):
                            if is_adult_func(dob) == "True":
                                row_data.append(dob)
                            else:
                                breaking = error = True
                                error_message = (
                                    "Parent at row number "
                                    + str(row_no + 1)
                                    + " is not an adult"
                                )
                            break
                        else:
                            breaking = error = True
                            error_message = (
                                "Invalid parent's 'Date of Birth' at row number "
                                + str(row_no + 1)
                                + ", please enter the input in the format given in the instructions"
                            )
                            break
            parent_data.append(row_data)

        if error == True:
            return render(
                request,
                "teacher/bulkregistration.html",
                {
                    "page_type": "bulk_register",
                    "my_messages": {"error": error_message},
                },
            )

        breaking = error = False
        error_message = ""
        student_data = []
        for row_no, row in enumerate(studentSheet.iter_rows()):
            if breaking == True:
                break
            if row_no == 0 or row_no == 1:
                continue
            row_data = []
            for cell in row:
                if cell.column_letter == "A":
                    if cell.value == None:
                        breaking = True
                        break
                    if not valid_name(cell.value):
                        breaking = error = True
                        error_message = (
                            "Invalid student's 'First Name' at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        row_data.append(str(cell.value))
                elif cell.column_letter == "B":
                    if cell.value == None:
                        row_data.append(None)
                    elif not valid_name(cell.value):
                        breaking = error = True
                        error_message = (
                            "Invalid student's 'Middle Name' at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        row_data.append(str(cell.value))
                elif cell.column_letter == "C":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Student's 'Last Name' missing at row number "
                            + str(row_no + 1)
                        )
                        break
                    elif not valid_name(cell.value):
                        breaking = error = True
                        error_message = (
                            "Invalid student's 'Last Name' at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        row_data.append(str(cell.value))
                elif cell.column_letter == "D":
                    if cell.value == None:
                        row_data.append(None)
                    elif not valid_aadhar(str(cell.value)):
                        breaking = error = True
                        error_message = (
                            "Invalid student's 'Aadhar number' at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        row_data.append(
                            str(cell.value)[:4]
                            + " "
                            + str(cell.value)[4:8]
                            + " "
                            + str(cell.value)[8:]
                        )
                elif cell.column_letter == "E":
                    if cell.value == None:
                        row_data.append(None)
                    elif not valid_email(cell.value):
                        breaking = error = True
                        error_message = (
                            "Invalid student's 'Email' at row number " + str(row_no + 1)
                        )
                        break
                    else:
                        row_data.append(str(cell.value))
                elif cell.column_letter == "F":
                    if cell.value == None:
                        row_data.append(None)
                    elif not valid_mobile_no(cell.value):
                        breaking = error = True
                        error_message = (
                            "Invalid student's 'Mobile Number' at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        row_data.append(str(cell.value))
                elif cell.column_letter == "G":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Student's 'Gender' missing at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        if check_gender(str(cell.value).capitalize()):
                            row_data.append(str(cell.value).capitalize())
                        else:
                            breaking = error = True
                            error_message = (
                                "Invalid 'Gender' input for 'Students Data' at row number  "
                                + str(row_no + 1)
                                + ", please select an option from the dropdown menu"
                            )
                            break
                elif cell.column_letter == "H":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Student's 'Date of Birth' missing at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        dob = (
                            str(cell.value)[8:10]
                            + "/"
                            + str(cell.value)[5:7]
                            + "/"
                            + str(cell.value)[:4]
                        )
                        if (
                            re.match("^[0-9]{2}/[0-9]{2}/[0-9]{4}$", dob)
                            or re.match("^[0-9]{2}-[0-9]{2}-[0-9]{4}$", dob)
                        ) and valid_date(dob):
                            if valid_dob(str(cell.value)[:10]):
                                row_data.append(dob)
                            else:
                                breaking = error = True
                                error_message = (
                                    "Student at row number "
                                    + str(row_no + 1)
                                    + " is not above the age of 5"
                                )
                                break
                        else:
                            breaking = error = True
                            error_message = (
                                "Invalid student's 'Date of Birth' at row number "
                                + str(row_no + 1)
                                + ", please enter the input in the format given in the instructions"
                            )
                            break
                elif cell.column_letter == "I":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Student's 'State' missing at row number " + str(row_no + 1)
                        )
                        break
                    else:
                        temp = check_state_city(True, 0, str(cell.value).capitalize())
                        if temp[0]:
                            row_data.append(str(cell.value).capitalize())
                            row_data.append(temp[1])
                        else:
                            breaking = error = True
                            error_message = (
                                "Invalid student's 'State' at row number "
                                + str(row_no + 1)
                            )
                            break
                elif cell.column_letter == "J":
                    if cell.value == None:
                        breaking = error = True
                        error_message = "Student's 'City' missing at row number " + str(
                            row_no + 1
                        )
                        break
                    else:
                        if check_state_city(
                            False, row_data[-1], str(cell.value).capitalize()
                        ):
                            row_data.pop()
                            row_data.append(str(cell.value).capitalize())
                        else:
                            breaking = error = True
                            error_message = (
                                "Invalid student's 'City' at row number "
                                + str(row_no + 1)
                            )
                            break
                elif cell.column_letter == "K":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Student's pincode missing at row number " + str(row_no + 1)
                        )
                        break
                    else:
                        if not valid_pincode(cell.value):
                            breaking = error = True
                            error_message = (
                                "Invalid student's pincode at row number "
                                + str(row_no + 1)
                            )
                            break
                        else:
                            row_data.append(str(cell.value))
                elif cell.column_letter == "L":
                    if cell.value == None:
                        breaking = error = True
                        error_message = (
                            "Student's 'Unique Number' missing at row number "
                            + str(row_no + 1)
                        )
                        break
                    else:
                        row_data.append(str(cell.value))
                elif cell.column_letter == "M":
                    dob = row_data[-5]
                    if cell.value == None:
                        if is_adult_func(dob) == "True":
                            row_data.append("ADULT")
                        else:
                            breaking = error = True
                            error_message = (
                                "Student at row number "
                                + str(row_no + 1)
                                + " is not an adult, please provide 'Parent's Ref ID' for the student"
                            )
                            break
                    else:
                        if is_adult_func(dob) == "True":
                            breaking = error = True
                            error_message = (
                                "Student at row number "
                                + str(row_no + 1)
                                + " is an adult, please remove 'Parent's Ref ID' for the student"
                            )
                            break
                        try:
                            parent_ref_id = int(cell.value)
                            if parent_ref_id <= len(parent_data):
                                row_data.append(str(cell.value))
                            else:
                                breaking = error = True
                                error_message = (
                                    "Invalid 'Parent's Ref ID' for student data at row number "
                                    + str(row_no + 1)
                                )
                                break
                        except:
                            breaking = error = True
                            error_message = (
                                "Invalid 'Parent's Ref ID' for student data at row number "
                                + str(row_no + 1)
                            )
                            break
            student_data.append(row_data)

        if error == True:
            return render(
                request,
                "teacher/bulkregistration.html",
                {
                    "page_type": "bulk_register",
                    "my_messages": {"error": error_message},
                },
            )

        parent_data.pop()
        student_data.pop()
        if len(student_data) == 0:
            return render(
                request,
                "teacher/bulkregistration.html",
                {
                    "page_type": "bulk_register",
                    "my_messages": {"error": "No student data found!"},
                },
            )

        our_user = TeacherInCharge.objects.filter(user=request.user).first()
        organization = our_user.organization
        parent_created = {}
        for i in student_data:
            password = random_password_generator()
            student_user = User.objects.create_user(
                username=username_generator(i[0], i[2]),
                password=password,
            )
            student_user.save()
            student_group = Group.objects.get(name="Students")
            student_user.groups.add(student_group)
            student_user.save()
            if i[-1] == "ADULT":
                StudentsInfo.objects.create(
                    user=student_user,
                    profile_pic="accounts/default.svg",
                    fname=encryptionHelper.encrypt(i[0]),
                    mname=encryptionHelper.encrypt(i[1]) if i[1] else None,
                    lname=encryptionHelper.encrypt(i[2]),
                    aadhar=encryptionHelper.encrypt(i[3]) if i[3] else None,
                    email=encryptionHelper.encrypt(i[4]) if i[4] else None,
                    mobile_no=encryptionHelper.encrypt(i[5]) if i[5] else None,
                    gender=encryptionHelper.encrypt(i[6]),
                    dob=encryptionHelper.encrypt(i[7]),
                    state=State.objects.get(state=i[8]),
                    city=City.objects.get(city=i[9]),
                    pincode=encryptionHelper.encrypt(i[10]),
                    unique_no=encryptionHelper.encrypt(i[11]),
                    organization=organization,
                    adult=encryptionHelper.encrypt("True"),
                    first_password=password,
                    password_changed=False,
                )
            else:
                if i[-1] in parent_created:
                    StudentsInfo.objects.create(
                        user=student_user,
                        profile_pic="accounts/default.svg",
                        fname=encryptionHelper.encrypt(i[0]),
                        mname=encryptionHelper.encrypt(i[1]) if i[1] else None,
                        lname=encryptionHelper.encrypt(i[2]),
                        aadhar=encryptionHelper.encrypt(i[3]) if i[3] else None,
                        email=encryptionHelper.encrypt(i[4]) if i[4] else None,
                        mobile_no=encryptionHelper.encrypt(i[5]) if i[5] else None,
                        gender=encryptionHelper.encrypt(i[6]),
                        dob=encryptionHelper.encrypt(i[7]),
                        state=State.objects.get(state=i[8]),
                        city=City.objects.get(city=i[9]),
                        pincode=encryptionHelper.encrypt(i[10]),
                        unique_no=encryptionHelper.encrypt(i[11]),
                        organization=organization,
                        adult=encryptionHelper.encrypt("False"),
                        parent=parent_created[i[-1]],
                        first_password=password,
                        password_changed=False,
                    )
                else:
                    this_parent_data = parent_data[int(i[-1]) - 1]
                    if this_parent_data[0]:
                        parent_user = User.objects.filter(
                            username=this_parent_data[0]
                        ).first()
                        parent = ParentsInfo.objects.filter(user=parent_user).first()
                        parent_created[i[-1]] = parent
                    else:
                        parent_password = random_password_generator()
                        parent_user = User.objects.create_user(
                            username=username_generator(
                                this_parent_data[1], this_parent_data[3]
                            ),
                            password=parent_password,
                        )
                        parent_user.save()
                        parent_group = Group.objects.get(name="Parents")
                        parent_user.groups.add(parent_group)
                        parent_user.save()
                        parent = ParentsInfo.objects.create(
                            user=parent_user,
                            profile_pic="accounts/default.svg",
                            fname=encryptionHelper.encrypt(this_parent_data[1]),
                            mname=encryptionHelper.encrypt(this_parent_data[2])
                            if this_parent_data[2]
                            else None,
                            lname=encryptionHelper.encrypt(this_parent_data[3]),
                            aadhar=encryptionHelper.encrypt(this_parent_data[4])
                            if this_parent_data[4]
                            else None,
                            email=encryptionHelper.encrypt(this_parent_data[5])
                            if this_parent_data[5]
                            else None,
                            mobile_no=encryptionHelper.encrypt(this_parent_data[6])
                            if this_parent_data[6]
                            else None,
                            gender=encryptionHelper.encrypt(this_parent_data[7]),
                            dob=encryptionHelper.encrypt(this_parent_data[8]),
                            first_password=parent_password,
                            password_changed=False,
                        )
                        parent_created[i[-1]] = parent
                    StudentsInfo.objects.create(
                        user=student_user,
                        profile_pic="accounts/default.svg",
                        fname=encryptionHelper.encrypt(i[0]),
                        mname=encryptionHelper.encrypt(i[1]) if i[1] else None,
                        lname=encryptionHelper.encrypt(i[2]),
                        aadhar=encryptionHelper.encrypt(i[3]) if i[3] else None,
                        email=encryptionHelper.encrypt(i[4]) if i[4] else None,
                        mobile_no=encryptionHelper.encrypt(i[5]) if i[5] else None,
                        gender=encryptionHelper.encrypt(i[6]),
                        dob=encryptionHelper.encrypt(i[7]),
                        state=State.objects.get(state=i[8]),
                        city=City.objects.get(city=i[9]),
                        pincode=encryptionHelper.encrypt(i[10]),
                        unique_no=encryptionHelper.encrypt(i[11]),
                        organization=organization,
                        adult=encryptionHelper.encrypt("False"),
                        parent=parent,
                        first_password=password,
                        password_changed=False,
                    )

        return render(
            request,
            "teacher/bulkregistration.html",
            {
                "page_type": "bulk_register",
                "my_messages": {
                    "success": "Bulk Registration Successful. Download the Login Credentials from the sidebar on the left."
                },
            },
        )


@login_required(login_url="accounts:loginlink")
@user_passes_test(
    lambda user: is_teacher(user),
    login_url="accounts:forbidden",
)
@password_change_required
def view_teacher_profile(request):
    if request.method == "GET":
        user = request.user
        if is_teacher(user):
            teacher = TeacherInCharge.objects.filter(user=user).first()

            teacher.fname = encryptionHelper.decrypt(teacher.fname)
            teacher.lname = encryptionHelper.decrypt(teacher.lname)

            if teacher.mname:
                teacher.mname = encryptionHelper.decrypt(teacher.mname)
                if teacher.mname == "":
                    teacher.mname = ""
            else:
                teacher.mname = ""

            if teacher.aadhar:
                teacher.aadhar = encryptionHelper.decrypt(teacher.aadhar)
                if teacher.aadhar == "":
                    teacher.aadhar = "-"
            else:
                teacher.aadhar = "-"

            if teacher.email:
                teacher.email = encryptionHelper.decrypt(teacher.email)
                if teacher.email == "":
                    teacher.email = "-"
            else:
                teacher.email = "-"

            if teacher.mobile_no:
                teacher.mobile_no = encryptionHelper.decrypt(teacher.mobile_no)
                if teacher.mobile_no == "":
                    teacher.mobile_no = "-"
            else:
                teacher.mobile_no = "-"

            teacher.dob = encryptionHelper.decrypt(teacher.dob)
            teacher.gender = encryptionHelper.decrypt(teacher.gender)

            return render(
                request,
                "teacher/view_teacher_profile.html",
                {"page_type": "view_teacher_profile", "teacher": teacher},
            )


@login_required(login_url="accounts:loginlink")
@user_passes_test(
    lambda user: is_teacher(user),
    login_url="accounts:forbidden",
)
@password_change_required
def edit_teacher_profile(request):
    teacher = TeacherInCharge.objects.filter(user=request.user).first()
    organization = teacher.organization
    if request.method == "GET":
        initial_dict = {
            "fname": encryptionHelper.decrypt(teacher.fname),
            "lname": encryptionHelper.decrypt(teacher.lname),
            "profile_pic": teacher.profile_pic,
            "dob": encryptionHelper.decrypt(teacher.dob),
            "gender": encryptionHelper.decrypt(teacher.gender),
        }

        if teacher.mname:
            initial_dict["mname"] = encryptionHelper.decrypt(teacher.mname)
        if teacher.aadhar:
            initial_dict["aadhar"] = encryptionHelper.decrypt(teacher.aadhar)
        if teacher.email:
            initial_dict["email"] = encryptionHelper.decrypt(teacher.email)
        if teacher.mobile_no:
            initial_dict["mobile_no"] = encryptionHelper.decrypt(teacher.mobile_no)

        form = TeachersInfoForm(request.POST or None, initial=initial_dict)
        form.fields["dob"].disabled = True

        return render(
            request,
            "teacher/update_teachers_info.html",
            {
                "form": form,
                "organization": organization,
            },
        )
    else:
        form = TeachersInfoForm(request.POST, request.FILES)
        form.fields["dob"].disabled = True
        form.fields["dob"].initial = encryptionHelper.decrypt(teacher.dob)

        if form.is_valid():
            teacher = TeacherInCharge.objects.filter(user=request.user).first()

            if teacher.mname:
                teacher.mname = encryptionHelper.encrypt(request.POST["mname"])
            if teacher.aadhar:
                teacher.aadhar = encryptionHelper.encrypt(request.POST["aadhar"])
            if teacher.email:
                teacher.email = encryptionHelper.encrypt(request.POST["email"])
            if teacher.mobile_no:
                teacher.mobile_no = encryptionHelper.encrypt(request.POST["mobile_no"])

            teacher.fname = encryptionHelper.encrypt(request.POST["fname"])
            teacher.lname = encryptionHelper.encrypt(request.POST["lname"])
            teacher.gender = encryptionHelper.encrypt(request.POST["gender"])

            if request.FILES:
                if request.FILES["profile_pic"].size > 5 * 1024 * 1024:
                    form.add_error("profile_pic", "File size must be less than 5MB.")

                    return render(
                        request,
                        "teacher/update_teachers_info.html",
                        {
                            "form": form,
                            "organization": organization,
                        },
                    )
                else:
                    x = teacher.profile_pic.url.split("/media/accounts/")
                    if x[1] != "default.svg":
                        file = settings.MEDIA_ROOT + "/accounts/" + x[1]
                        os.remove(file)
                    teacher.profile_pic = request.FILES["profile_pic"]
            else:
                if "profile_pic-clear" in request.POST.keys():
                    x = teacher.profile_pic.url.split("/media/accounts/")
                    if x[1] != "default.svg":
                        file = settings.MEDIA_ROOT + "/accounts/" + x[1]
                        os.remove(file)
                    teacher.profile_pic = "accounts/default.svg"

            teacher.save()
            return redirect("accounts:view_teacher_profile")
        else:

            return render(
                request,
                "teacher/update_teachers_info.html",
                {"form": form, "organization": organization},
            )
