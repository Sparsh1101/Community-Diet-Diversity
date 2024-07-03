import io
import openpyxl
import xlsxwriter
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.models import Group, User
from django.contrib.auth.decorators import login_required, user_passes_test
from .decorators import *
from .models import *
from .forms import *
from .helper_functions import *
from shared.encryption import EncryptionHelper
from django.utils import timezone
from django.conf import settings

encryptionHelper = EncryptionHelper()


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def coordinator_dashboard(request):
    teachers = (
        CoordinatorInCharge.objects.filter(user=request.user)
        .first()
        .teacherincharge_set.all()
    )
    for teacher in teachers:
        teacher.fname = encryptionHelper.decrypt(teacher.fname)
        teacher.lname = encryptionHelper.decrypt(teacher.lname)
        if teacher.mobile_no:
            teacher.mobile_no = encryptionHelper.decrypt(teacher.mobile_no)
            if teacher.mobile_no == "":
                teacher.mobile_no = "-"
        else:
            teacher.mobile_no = "-"
        if teacher.email:
            teacher.email = encryptionHelper.decrypt(teacher.email)
            if teacher.email == "":
                teacher.email = "-"
        else:
            teacher.email = "-"
    if "my_messages" in request.session:
        my_messages = request.session["my_messages"]
        del request.session["my_messages"]
        return render(
            request,
            "coordinator/all_teachers.html",
            {
                "teachers": teachers,
                "page_type": "coordinator_dashboard",
                "my_messages": my_messages,
            },
        )
    return render(
        request,
        "coordinator/all_teachers.html",
        {"teachers": teachers, "page_type": "coordinator_dashboard"},
    )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def switchTeachersList(request, id, teacher_id):
    if request.method == "GET":
        session = Session.objects.filter(id=id).first()
        objects_in_sessions = Teacher_Session.objects.filter(session=session)
        teachers_in_sessions = []
        teacher = TeacherInCharge.objects.filter(id=teacher_id).first()
        coordinator = CoordinatorInCharge.objects.filter(user=request.user).first()
        organization = coordinator.organization
        for object1 in objects_in_sessions:
            if object1.teacher != teacher:
                teachers_in_sessions.append(object1.teacher)

        for teacher in teachers_in_sessions:
            teacher.fname = encryptionHelper.decrypt(teacher.fname)
            teacher.lname = encryptionHelper.decrypt(teacher.lname)
        return render(
            request,
            "coordinator/switch_teachers_list.html",
            {"page_type": "switch_teachers_list", "teachers": teachers_in_sessions},
        )
    else:
        new_teacher_id = request.POST.get("new_teacher")
        new_teacher = TeacherInCharge.objects.filter(id=new_teacher_id).first()
        session = Session.objects.filter(id=id).first()
        teacher = TeacherInCharge.objects.filter(id=teacher_id).first()
        teacher_session = Teacher_Session.objects.filter(
            session=session, teacher=teacher
        ).delete()
        students = StudentsInfo.objects.filter(session=session, teacher=teacher)
        for student in students:
            student.teacher = new_teacher
            student_session = Student_Session.objects.filter(
                session=session, student=student, teacher=teacher
            ).first()
            student_session.teacher = new_teacher
            student.save()
            student_session.save()

        forms = FormDetails.objects.filter(teacher=teacher)
        for form in forms:
            form.teacher = new_teacher
            form.save()

        return redirect("accounts:view_session_teachers", id, 1)


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def addTeacherForm(request):
    if request.method == "GET":
        form = TeachersInfoForm()
        return render(
            request,
            "coordinator/add_teacher.html",
            {"form": form},
        )
    else:
        form = TeachersInfoForm(request.POST)
        if form.is_valid():
            password = random_password_generator()
            teacheruser = User.objects.create_user(
                username=username_generator(
                    request.POST["fname"], request.POST["lname"]
                ),
                password=password,
            )
            teacheruser.save()
            teacher_group = Group.objects.get(name="Teachers")
            teacheruser.groups.add(teacher_group)
            teacheruser.save()
            teacher = form.save(commit=False)
            teacher.user = teacheruser
            teacher.email = encryptionHelper.encrypt(request.POST["email"])
            teacher.fname = encryptionHelper.encrypt(request.POST["fname"])
            teacher.mname = encryptionHelper.encrypt(request.POST["mname"])
            teacher.lname = encryptionHelper.encrypt(request.POST["lname"])
            teacher.aadhar = encryptionHelper.encrypt(request.POST["aadhar"])
            teacher.dob = encryptionHelper.encrypt(request.POST["dob"])
            teacher.gender = encryptionHelper.encrypt(request.POST["gender"])
            teacher.mobile_no = encryptionHelper.encrypt(request.POST["mobile_no"])
            teacher.organization = Organization.objects.filter(
                name=CoordinatorInCharge.objects.filter(user=request.user)
                .first()
                .organization
            ).first()
            teacher.coordinator = CoordinatorInCharge.objects.filter(
                user=request.user
            ).first()
            teacher.profile_pic = "accounts/default.svg"
            teacher.password_changed = False
            teacher.first_password = password
            teacher.save()
            request.session["my_messages"] = {
                "success": "Registration Successful. Download the Login Credentials from the sidebar on the left."
            }
            return redirect("accounts:coordinator_dashboard")
        else:
            return render(
                request,
                "coordinator/add_teacher.html",
                {"form": form},
            )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def allSessions(request):
    sessions = (
        CoordinatorInCharge.objects.filter(user=request.user).first().session_set.all()
    )
    open_sessions = []
    close_sessions = []
    upcoming_sessions = []
    open = False
    close = False
    upcoming = False
    for session in sessions:
        if session.start_date > timezone.now():
            upcoming_sessions.append(session)
            upcoming = True
        elif session.end_date == None:
            open_sessions.append(session)
            open = True
        else:
            close_sessions.append(session)
            close = True
    return render(
        request,
        "coordinator/all_sessions.html",
        {
            "open_sessions": open_sessions,
            "upcoming_sessions": upcoming_sessions,
            "close_sessions": close_sessions,
            "open": open,
            "upcoming": upcoming,
            "close": close,
            "page_type": "all_sessions",
        },
    )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def addSessionForm(request):
    if request.method == "GET":
        form = SessionsInfoForm()
        return render(
            request,
            "coordinator/add_session.html",
            {"form": form},
        )
    else:
        form = SessionsInfoForm(request.POST)
        if form.is_valid():
            session = form.save(commit=False)
            session.coordinator = CoordinatorInCharge.objects.filter(
                user=request.user
            ).first()
            session.save()
            return redirect("accounts:all_sessions")
        else:
            return render(
                request,
                "coordinator/add_session.html",
                {"form": form},
            )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def viewSessionTeachers(request, id, open_id):
    session = Session.objects.filter(id=id).first()
    objects = Teacher_Session.objects.filter(session=session)
    teachers = []
    for object in objects:
        object.teacher.fname = encryptionHelper.decrypt(object.teacher.fname)
        object.teacher.lname = encryptionHelper.decrypt(object.teacher.lname)
        if object.teacher.mobile_no:
            object.teacher.mobile_no = encryptionHelper.decrypt(
                object.teacher.mobile_no
            )
            if object.teacher.mobile_no == "":
                object.teacher.mobile_no = "-"
        else:
            object.teacher.mobile_no = "-"
        if object.teacher.email:
            object.teacher.email = encryptionHelper.decrypt(object.teacher.email)
            if object.teacher.email == "":
                object.teacher.email = "-"
        else:
            object.teacher.email = "-"
        teachers.append(object.teacher)
    if "my_messages" in request.session:
        my_messages = request.session["my_messages"]
        del request.session["my_messages"]
        return render(
            request,
            "coordinator/view_session_teachers.html",
            {
                "teachers": teachers,
                "session": session,
                "my_messages": my_messages,
                "open_id": open_id,
                "page_type": "view_session_teachers",
            },
        )
    else:
        return render(
            request,
            "coordinator/view_session_teachers.html",
            {
                "teachers": teachers,
                "session": session,
                "open_id": open_id,
                "page_type": "view_session_teachers",
            },
        )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def getSessionTeachersTemplate(request):
    output = io.BytesIO()
    wb = xlsxwriter.Workbook(output)
    ws = wb.add_worksheet("Session Teachers Data")
    columns = [
        "Teacher Username",
    ]
    for col_num in range(len(columns)):
        ws.write(0, col_num, columns[col_num])
    wb.close()
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=sessionTeacherTemplate.xlsx"
    return response


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def addSessionTeachers(request, id):
    if request.method == "GET":
        return render(
            request,
            "coordinator/add_session_teachers.html",
            {"page_type": "add_session_teachers"},
        )
    else:
        try:
            excel_file = request.FILES["excel_file"]
            if excel_file.name[-4:] == "xlsx":
                try:
                    wb = openpyxl.load_workbook(excel_file)
                    teacherSheet = wb["Session Teachers Data"]
                except:
                    return render(
                        request,
                        "coordinator/add_session_teachers.html",
                        {
                            "page_type": "add_session_teachers",
                            "my_messages": {
                                "error": "Incorrect file uploaded, please use the template provided above."
                            },
                        },
                    )
            else:
                return render(
                    request,
                    "coordinator/add_session_teachers.html",
                    {
                        "page_type": "add_session_teachers",
                        "my_messages": {
                            "error": "Incorrect file type, only .xlsx files are allowed!"
                        },
                    },
                )
        except:
            return render(
                request,
                "coordinator/add_session_teachers.html",
                {
                    "page_type": "add_session_teachers",
                    "my_messages": {"error": "Sorry something went wrong!"},
                },
            )

        organization = (
            CoordinatorInCharge.objects.filter(user=request.user).first().organization
        )
        breaking = error = False
        error_message = ""
        teacher_data = []
        for row_no, row in enumerate(teacherSheet.iter_rows()):
            if breaking == True:
                break
            if row_no == 0:
                continue
            for cell in row:
                if cell.column_letter == "A":
                    if cell.value == None:
                        breaking = True
                        break
                    else:
                        if User.objects.filter(username=cell.value).exists():
                            user = User.objects.filter(username=cell.value).first()
                            if is_teacher(user):
                                teacher_user = TeacherInCharge.objects.filter(
                                    user=user
                                ).first()
                                if teacher_user.organization == organization:
                                    teacher_data.append(teacher_user)
                                else:
                                    breaking = error = True
                                    error_message = (
                                        "Teacher at row number "
                                        + str(row_no + 1)
                                        + " does not belong to your organization"
                                    )
                                    break
                            else:
                                breaking = error = True
                                error_message = (
                                    "User at row number "
                                    + str(row_no + 1)
                                    + " is not a teacher"
                                )
                                break
                        else:
                            breaking = error = True
                            error_message = "Invalid username at row number " + str(
                                row_no + 1
                            )
                            break

        if error == True:
            return render(
                request,
                "coordinator/add_session_teachers.html",
                {
                    "page_type": "add_session_teachers",
                    "my_messages": {"error": error_message},
                },
            )

        session = Session.objects.filter(id=id).first()
        teacher_session = Teacher_Session.objects.filter(session=session)
        teacher_session_teacher_objects = []
        for i in teacher_session:
            teacher_session_teacher_objects.append(i.teacher)
        teacher_added = 0
        teacher_exists = 0
        for teacher_user in teacher_data:
            if teacher_user in teacher_session_teacher_objects:
                teacher_exists += 1
            else:
                new_teacher_session = Teacher_Session()
                new_teacher_session.session = session
                new_teacher_session.teacher = teacher_user
                new_teacher_session.save()
                teacher_added += 1

        if teacher_added == 0:
            my_messages = {
                "success": "Registration successful. Already registered: "
                + str(teacher_exists)
            }
        elif teacher_exists == 0:
            my_messages = {
                "success": "Registration successful. Newly registered: "
                + str(teacher_added)
            }
        else:
            my_messages = {
                "success": "Registration successful. Already registered: "
                + str(teacher_exists)
                + ", Newly registered: "
                + str(teacher_added)
            }
        request.session["my_messages"] = my_messages
        return redirect("accounts:view_session_teachers", id, 1)


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def addSessionTeachersList(request, id):
    if request.method == "GET":
        session = Session.objects.filter(id=id).first()
        teachers_in_sessions = Teacher_Session.objects.filter(session=session)
        teachers_in_sessions_id = []
        coordinator = CoordinatorInCharge.objects.filter(user=request.user).first()
        organization = coordinator.organization
        for teacher in teachers_in_sessions:
            teachers_in_sessions_id.append(teacher.teacher.id)

        teachers = TeacherInCharge.objects.filter(
            coordinator=coordinator, organization=organization
        ).exclude(id__in=teachers_in_sessions_id)

        for teacher in teachers:
            teacher.fname = encryptionHelper.decrypt(teacher.fname)
            teacher.lname = encryptionHelper.decrypt(teacher.lname)
        return render(
            request,
            "coordinator/add_session_teachers_list.html",
            {"page_type": "add_session_teachers_list", "teachers": teachers},
        )
    else:
        teachers_id_list = request.POST.getlist("chk")
        session = Session.objects.filter(id=id).first()
        for t_id in teachers_id_list:
            teacher = TeacherInCharge.objects.filter(id=t_id).first()
            teacher_session = Teacher_Session()
            teacher_session.session = session
            teacher_session.teacher = teacher
            teacher_session.save()

        return redirect("accounts:view_session_teachers", id, 1)


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def removeSessionTeacher(request, session_id, teacher_id):
    session = Session.objects.filter(id=session_id).first()
    teacher = TeacherInCharge.objects.filter(id=teacher_id).first()
    students = StudentsInfo.objects.filter(teacher=teacher, session=session)
    for student in students:
        student.teacher = None
        student.session = None
        student.save()
    Teacher_Session.objects.filter(teacher=teacher).delete()
    student_session = Student_Session.objects.filter(
        session=session, teacher=teacher
    ).delete()
    request.session["my_messages"] = {
        "success": "Teacher removed from session successfully"
    }
    return redirect("accounts:view_session_teachers", session_id, 1)


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def coordinator_reset_password(request):
    if request.method == "GET":
        form = CoordPasswordReset()
        return render(
            request,
            "coordinator/coordinator_reset_password.html",
            {"form": form, "page_type": "reset_password"},
        )
    else:
        form = CoordPasswordReset(request.POST)
        organization = CoordinatorInCharge.objects.get(user=request.user).organization
        if form.is_valid():
            username = form.cleaned_data["username"]
            if User.objects.filter(username=username).exists():
                user = User.objects.get(username=username)
                if is_teacher(user):
                    user_teacher = TeacherInCharge.objects.get(user=user)
                    if user_teacher.organization == organization:
                        password = random_password_generator()
                        user_teacher.first_password = password
                        user_teacher.password_changed = False
                        user_teacher.save()
                        user.set_password(password)
                        user.save()
                        return render(
                            request,
                            "coordinator/coordinator_reset_password.html",
                            {
                                "form": form,
                                "page_type": "reset_password",
                                "my_messages": {
                                    "success": "Password reset successfull. Download the Login Credentials from the sidebar on the left."
                                },
                            },
                        )
                    else:
                        return render(
                            request,
                            "coordinator/coordinator_reset_password.html",
                            {
                                "form": form,
                                "page_type": "reset_password",
                                "my_messages": {
                                    "error": "Teacher does not belong to your organization."
                                },
                            },
                        )
                elif is_parent(user):
                    user_parent = ParentsInfo.objects.get(user=user)
                    if user_parent.organization == organization:
                        password = random_password_generator()
                        user_parent.first_password = password
                        user_parent.password_changed = False
                        user_parent.save()
                        user.set_password(password)
                        user.save()
                        return render(
                            request,
                            "coordinator/coordinator_reset_password.html",
                            {
                                "form": form,
                                "page_type": "reset_password",
                                "my_messages": {
                                    "success": "Password reset successfull. Download the Login Credentials from the sidebar on the left."
                                },
                            },
                        )
                    else:
                        return render(
                            request,
                            "coordinator/coordinator_reset_password.html",
                            {
                                "form": form,
                                "page_type": "reset_password",
                                "my_messages": {
                                    "error": "Parent does not belong to your organization."
                                },
                            },
                        )
                elif is_student(user):
                    user_student = StudentsInfo.objects.get(user=user)
                    if user_student.organization == organization:
                        password = random_password_generator()
                        user_student.first_password = password
                        user_student.password_changed = False
                        user_student.save()
                        user.set_password(password)
                        user.save()
                        return render(
                            request,
                            "coordinator/coordinator_reset_password.html",
                            {
                                "form": form,
                                "page_type": "reset_password",
                                "my_messages": {
                                    "success": "Password reset successfull. Download the Login Credentials from the sidebar on the left."
                                },
                            },
                        )
                    else:
                        return render(
                            request,
                            "coordinator/coordinator_reset_password.html",
                            {
                                "form": form,
                                "page_type": "reset_password",
                                "my_messages": {
                                    "error": "Student does not belong to your organization."
                                },
                            },
                        )
                else:
                    return render(
                        request,
                        "coordinator/coordinator_reset_password.html",
                        {
                            "form": form,
                            "page_type": "reset_password",
                            "my_messages": {
                                "error": "Provided user is not a teacher/parent/student."
                            },
                        },
                    )
            else:
                return render(
                    request,
                    "coordinator/coordinator_reset_password.html",
                    {
                        "form": form,
                        "page_type": "reset_password",
                        "my_messages": {"error": "Invalid username."},
                    },
                )
        else:
            return render(
                request,
                "coordinator/coordinator_reset_password.html",
                {"form": form, "page_type": "reset_password"},
            )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def teachers_data_download(request):
    output = io.BytesIO()
    wb = xlsxwriter.Workbook(output)
    bold = wb.add_format({"bold": True})
    merge_format = wb.add_format(
        {
            "bold": True,
            "align": "center",
            "valign": "vcenter",
        }
    )
    credentials = wb.add_worksheet("Teachers Data")
    credentials.set_row(0, 40)
    credentials.merge_range(
        "A1:J1",
        '"***" indicates password changed by the user. (MAKE SURE THAT THE FOLLOWING SHEET CANNOT BE ACCESSED BY UNAUTHORIZED USERS SINCE IT CONTAINS SENSITIVE DATA)',
        merge_format,
    )
    columns = [
        "USERNAME",
        "PASSWORD",
        "First Name",
        "Middle Name",
        "Last Name",
        "Aadhar Number",
        "Email",
        "Mobile Number",
        "Date of Birth",
        "Gender",
    ]
    credentials.set_column(0, len(columns) - 1, 18)
    for col_num in range(len(columns)):
        credentials.write(1, col_num, columns[col_num], bold)
    teacher = TeacherInCharge.objects.filter(
        organization=CoordinatorInCharge.objects.filter(user=request.user)
        .first()
        .organization,
    )
    for row_no, x in enumerate(teacher):
        credentials.write(row_no + 2, 0, x.user.username)
        if x.password_changed:
            credentials.write(row_no + 2, 1, "***")
        else:
            credentials.write(row_no + 2, 1, x.first_password)
        credentials.write(row_no + 2, 2, encryptionHelper.decrypt(x.fname))
        if x.mname:
            x.mname = encryptionHelper.decrypt(x.mname)
            if x.mname == "":
                x.mname = "-"
        else:
            x.mname = "-"
        credentials.write(row_no + 2, 3, x.mname)
        credentials.write(row_no + 2, 4, encryptionHelper.decrypt(x.lname))
        if x.aadhar:
            x.aadhar = encryptionHelper.decrypt(x.aadhar)
            if x.aadhar == "":
                x.aadhar = "-"
        else:
            x.aadhar = "-"
        credentials.write(row_no + 2, 5, x.aadhar)
        if x.email:
            x.email = encryptionHelper.decrypt(x.email)
            if x.email == "":
                x.email = "-"
        else:
            x.email = "-"
        credentials.write(row_no + 2, 6, x.email)
        if x.mobile_no:
            x.mobile_no = encryptionHelper.decrypt(x.mobile_no)
            if x.mobile_no == "":
                x.mobile_no = "-"
        else:
            x.mobile_no = "-"
        credentials.write(row_no + 2, 7, x.mobile_no)
        dob = encryptionHelper.decrypt(x.dob).split("/")
        dob[1], dob[0] = dob[0], dob[1]
        dob = "/".join(dob)
        credentials.write(row_no + 2, 8, dob)
        credentials.write(row_no + 2, 9, encryptionHelper.decrypt(x.gender))
    wb.close()
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=Teachers Data.xlsx"
    return response


@login_required(login_url="accounts:loginlink")
@user_passes_test(
    lambda user: is_coordinator(user) or is_teacher(user),
    login_url="accounts:forbidden",
)
@password_change_required
def parents_and_students_data_download(request):
    output = io.BytesIO()
    wb = xlsxwriter.Workbook(output)
    bold = wb.add_format({"bold": True})
    merge_format = wb.add_format(
        {
            "bold": True,
            "align": "center",
            "valign": "vcenter",
        }
    )
    parent = wb.add_worksheet("Parents Data")
    parent.set_row(0, 40)
    parent.merge_range(
        "A1:K1",
        '"***" indicates password changed by the user. (MAKE SURE THAT THE FOLLOWING SHEET CANNOT BE ACCESSED BY UNAUTHORIZED USERS SINCE IT CONTAINS SENSITIVE DATA)',
        merge_format,
    )
    columns = [
        "USERNAME",
        "PASSWORD",
        "Ref ID",
        "First Name",
        "Middle Name",
        "Last Name",
        "Aadhar Number",
        "Email",
        "Mobile Number",
        "Date of Birth",
        "Gender",
    ]
    parent.set_column(0, len(columns) - 1, 18)
    for col_num in range(len(columns)):
        parent.write(1, col_num, columns[col_num], bold)
    student = wb.add_worksheet("Students Data")
    student.set_row(0, 40)
    student.merge_range(
        "A1:V1",
        '"***" indicates password changed by the user. "Parent\'s Ref ID" indicates the parent registered along with the student and "-" indicates that the student is an adult (MAKE SURE THAT THE FOLLOWING SHEET CANNOT BE ACCESSED BY UNAUTHORIZED USERS SINCE IT CONTAINS SENSITIVE DATA)',
        merge_format,
    )
    columns = [
        "USERNAME",
        "PASSWORD",
        "Parent's Ref ID",
        "First Name",
        "Middle Name",
        "Last Name",
        "Aadhar Number",
        "Email",
        "Mobile Number",
        "Date of Birth",
        "Gender",
        "State",
        "City",
        "Pincode",
        "Unique Number",
        "Parent's Occupation",
        "Parent's Education",
        "Religious Belief",
        "Type of Family",
        "No. of Family Members",
        "Total Family Income",
        "Ration Card Color",
    ]
    student.set_column(0, len(columns) - 1, 18)
    for col_num in range(len(columns)):
        student.write(1, col_num, columns[col_num], bold)
    if is_coordinator(request.user):
        students = StudentsInfo.objects.filter(
            organization=CoordinatorInCharge.objects.filter(user=request.user)
            .first()
            .organization,
        )
    else:
        students = StudentsInfo.objects.filter(
            organization=TeacherInCharge.objects.filter(user=request.user)
            .first()
            .organization,
        )
    parents_list = []
    ref_count = 0
    for row_no, x in enumerate(students):
        student.write(row_no + 2, 0, x.user.username)
        if x.password_changed:
            student.write(row_no + 2, 1, "***")
        else:
            student.write(row_no + 2, 1, x.first_password)
        if encryptionHelper.decrypt(x.adult) == "True":
            student.write(row_no + 2, 2, "-")
        else:
            if x.parent in parents_list:
                for index, i in enumerate(parents_list):
                    if i == x.parent:
                        student.write(row_no + 2, 2, index + 1)
                        break
            else:
                ref_count += 1
                student.write(row_no + 2, 2, ref_count)
                parent_here = x.parent
                parents_list.append(parent_here)
                parent.write(ref_count + 1, 0, parent_here.user.username)
                if parent_here.password_changed:
                    parent.write(ref_count + 1, 1, "***")
                else:
                    parent.write(ref_count + 1, 1, parent_here.first_password)
                parent.write(ref_count + 1, 2, ref_count)
                parent.write(
                    ref_count + 1, 3, encryptionHelper.decrypt(parent_here.fname)
                )
                if parent_here.mname:
                    parent_here.mname = encryptionHelper.decrypt(parent_here.mname)
                    if parent_here.mname == "":
                        parent_here.mname = "-"
                else:
                    parent_here.mname = "-"
                parent.write(ref_count + 1, 4, parent_here.mname)
                parent.write(
                    ref_count + 1, 5, encryptionHelper.decrypt(parent_here.lname)
                )
                if parent_here.aadhar:
                    parent_here.aadhar = encryptionHelper.decrypt(parent_here.aadhar)
                    if parent_here.aadhar == "":
                        parent_here.aadhar = "-"
                else:
                    parent_here.aadhar = "-"
                parent.write(ref_count + 1, 6, parent_here.aadhar)
                if parent_here.email:
                    parent_here.email = encryptionHelper.decrypt(parent_here.email)
                    if parent_here.email == "":
                        parent_here.email = "-"
                else:
                    parent_here.email = "-"
                parent.write(ref_count + 1, 7, parent_here.email)
                if parent_here.mobile_no:
                    parent_here.mobile_no = encryptionHelper.decrypt(
                        parent_here.mobile_no
                    )
                    if parent_here.mobile_no == "":
                        parent_here.mobile_no = "-"
                else:
                    parent_here.mobile_no = "-"
                parent.write(ref_count + 1, 8, parent_here.mobile_no)
                dob = encryptionHelper.decrypt(parent_here.dob).split("/")
                dob[1], dob[0] = dob[0], dob[1]
                dob = "/".join(dob)
                parent.write(ref_count + 1, 9, dob)
                parent.write(
                    ref_count + 1, 10, encryptionHelper.decrypt(parent_here.gender)
                )
        student.write(row_no + 2, 3, encryptionHelper.decrypt(x.fname))
        if x.mname:
            x.mname = encryptionHelper.decrypt(x.mname)
            if x.mname == "":
                x.mname = "-"
        else:
            x.mname = "-"
        student.write(row_no + 2, 4, x.mname)
        student.write(row_no + 2, 5, encryptionHelper.decrypt(x.lname))
        if x.aadhar:
            x.aadhar = encryptionHelper.decrypt(x.aadhar)
            if x.aadhar == "":
                x.aadhar = "-"
        else:
            x.aadhar = "-"
        student.write(row_no + 2, 6, x.aadhar)
        if x.email:
            x.email = encryptionHelper.decrypt(x.email)
            if x.email == "":
                x.email = "-"
        else:
            x.email = "-"
        student.write(row_no + 2, 7, x.email)
        if x.mobile_no:
            x.mobile_no = encryptionHelper.decrypt(x.mobile_no)
            if x.mobile_no == "":
                x.mobile_no = "-"
        else:
            x.mobile_no = "-"
        student.write(row_no + 2, 8, x.mobile_no)
        dob = encryptionHelper.decrypt(x.dob).split("/")
        dob[1], dob[0] = dob[0], dob[1]
        dob = "/".join(dob)
        student.write(row_no + 2, 9, dob)
        student.write(row_no + 2, 10, encryptionHelper.decrypt(x.gender))
        student.write(row_no + 2, 11, x.state.state)
        student.write(row_no + 2, 12, x.city.city)
        student.write(row_no + 2, 13, encryptionHelper.decrypt(x.pincode))
        student.write(row_no + 2, 14, encryptionHelper.decrypt(x.unique_no))
        if x.secondary_reg:
            if not x.secondary_reg.occupation:
                occupation = "-"
            else:
                occupation = x.secondary_reg.occupation.occupation
            student.write(row_no + 2, 15, occupation)
            if not x.secondary_reg.edu:
                edu = "-"
            else:
                edu = x.secondary_reg.edu.education
            student.write(row_no + 2, 16, edu)
            if not x.secondary_reg.religion:
                religion = "-"
            else:
                religion = x.secondary_reg.religion.religion
            student.write(row_no + 2, 17, religion)
            if not x.secondary_reg.type_of_family:
                type_of_family = "-"
            else:
                type_of_family = x.secondary_reg.type_of_family.family
            student.write(row_no + 2, 18, type_of_family)
            if not x.secondary_reg.no_of_family_members:
                x.secondary_reg.no_of_family_members = "-"
            student.write(row_no + 2, 19, x.secondary_reg.no_of_family_members)
            if not x.secondary_reg.family_income:
                family_income = "-"
            else:
                family_income = x.secondary_reg.family_income.family_income
            student.write(row_no + 2, 20, family_income)
            if not x.secondary_reg.ration_card_color:
                ration_card_color = "-"
            else:
                ration_card_color = x.secondary_reg.ration_card_color.ration_card_color
            student.write(row_no + 2, 21, ration_card_color)
        else:
            student.write(row_no + 2, 15, "-")
            student.write(row_no + 2, 16, "-")
            student.write(row_no + 2, 17, "-")
            student.write(row_no + 2, 18, "-")
            student.write(row_no + 2, 19, "-")
            student.write(row_no + 2, 20, "-")
            student.write(row_no + 2, 21, "-")
    wb.close()
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response[
        "Content-Disposition"
    ] = "attachment; filename=Parents and Students Data.xlsx"
    return response


@login_required(login_url="accounts:loginlink")
@user_passes_test(
    lambda user: is_coordinator(user),
    login_url="accounts:forbidden",
)
@password_change_required
def view_coordinator_profile(request):
    if request.method == "GET":
        user = request.user
        if is_coordinator(user):
            coordinator = CoordinatorInCharge.objects.filter(user=user).first()

            coordinator.fname = encryptionHelper.decrypt(coordinator.fname)
            coordinator.lname = encryptionHelper.decrypt(coordinator.lname)

            if coordinator.mname:
                coordinator.mname = encryptionHelper.decrypt(coordinator.mname)
                if coordinator.mname == "":
                    coordinator.mname = ""
            else:
                coordinator.mname = ""

            if coordinator.aadhar:
                coordinator.aadhar = encryptionHelper.decrypt(coordinator.aadhar)
                if coordinator.aadhar == "":
                    coordinator.aadhar = "-"
            else:
                coordinator.aadhar = "-"

            if coordinator.email:
                coordinator.email = encryptionHelper.decrypt(coordinator.email)
                if coordinator.email == "":
                    coordinator.email = "-"
            else:
                coordinator.email = "-"

            if coordinator.mobile_no:
                coordinator.mobile_no = encryptionHelper.decrypt(coordinator.mobile_no)
                if coordinator.mobile_no == "":
                    coordinator.mobile_no = "-"
            else:
                coordinator.mobile_no = "-"

            coordinator.dob = encryptionHelper.decrypt(coordinator.dob)
            coordinator.gender = encryptionHelper.decrypt(coordinator.gender)

            return render(
                request,
                "coordinator/view_coordinator_profile.html",
                {"page_type": "view_coordinator_profile", "coordinator": coordinator},
            )


@login_required(login_url="accounts:loginlink")
@user_passes_test(
    lambda user: is_coordinator(user),
    login_url="accounts:forbidden",
)
@password_change_required
def edit_coordinator_profile(request):
    coordinator = CoordinatorInCharge.objects.filter(user=request.user).first()
    organization = coordinator.organization
    if request.method == "GET":
        initial_dict = {
            "fname": encryptionHelper.decrypt(coordinator.fname),
            "lname": encryptionHelper.decrypt(coordinator.lname),
            "profile_pic": coordinator.profile_pic,
            "dob": encryptionHelper.decrypt(coordinator.dob),
            "gender": encryptionHelper.decrypt(coordinator.gender),
        }

        if coordinator.mname:
            initial_dict["mname"] = encryptionHelper.decrypt(coordinator.mname)
        if coordinator.aadhar:
            initial_dict["aadhar"] = encryptionHelper.decrypt(coordinator.aadhar)
        if coordinator.email:
            initial_dict["email"] = encryptionHelper.decrypt(coordinator.email)
        if coordinator.mobile_no:
            initial_dict["mobile_no"] = encryptionHelper.decrypt(coordinator.mobile_no)

        form = CoordinatorsInfoForm(request.POST or None, initial=initial_dict)
        form.fields["dob"].disabled = True

        return render(
            request,
            "coordinator/update_coordinators_info.html",
            {"form": form, "organization": organization},
        )
    else:
        form = CoordinatorsInfoForm(request.POST, request.FILES)
        form.fields["dob"].disabled = True
        form.fields["dob"].initial = encryptionHelper.decrypt(coordinator.dob)

        if form.is_valid():
            coordinator = CoordinatorInCharge.objects.filter(user=request.user).first()

            if coordinator.mname:
                coordinator.mname = encryptionHelper.encrypt(request.POST["mname"])
            if coordinator.aadhar:
                coordinator.aadhar = encryptionHelper.encrypt(request.POST["aadhar"])
            if coordinator.email:
                coordinator.email = encryptionHelper.encrypt(request.POST["email"])
            if coordinator.mobile_no:
                coordinator.mobile_no = encryptionHelper.encrypt(
                    request.POST["mobile_no"]
                )

            coordinator.fname = encryptionHelper.encrypt(request.POST["fname"])
            coordinator.lname = encryptionHelper.encrypt(request.POST["lname"])
            coordinator.gender = encryptionHelper.encrypt(request.POST["gender"])

            if request.FILES:
                if request.FILES["profile_pic"].size > 5 * 1024 * 1024:
                    form.add_error("profile_pic", "File size must be less than 5MB.")

                    return render(
                        request,
                        "coordinator/update_coordinators_info.html",
                        {
                            "form": form,
                            "organization": organization,
                        },
                    )
                else:
                    x = coordinator.profile_pic.url.split("/media/accounts/")
                    if x[1] != "default.svg":
                        file = settings.MEDIA_ROOT + "/accounts/" + x[1]
                        os.remove(file)
                    coordinator.profile_pic = request.FILES["profile_pic"]
            else:
                if "profile_pic-clear" in request.POST.keys():
                    x = coordinator.profile_pic.url.split("/media/accounts/")
                    if x[1] != "default.svg":
                        file = settings.MEDIA_ROOT + "/accounts/" + x[1]
                        os.remove(file)
                    coordinator.profile_pic = "accounts/default.svg"

            coordinator.save()
            return redirect("accounts:view_coordinator_profile")
        else:
            return render(
                request,
                "coordinator/update_coordinators_info.html",
                {
                    "form": form,
                    "organization": organization,
                },
            )


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def switchTeachersUserList(request, teacher_id):
    og_teacher = TeacherInCharge.objects.filter(id=teacher_id).first()
    og_teacher_user = og_teacher.user
    organization = og_teacher.organization
    if request.method == "GET":
        teachers = organization.teacherincharge_set.all()
        teachers_without_og = []

        for teacher in teachers:
            if teacher != og_teacher:
                teachers_without_og.append(teacher)

        for teacher in teachers_without_og:
            teacher.fname = encryptionHelper.decrypt(teacher.fname)
            teacher.lname = encryptionHelper.decrypt(teacher.lname)
        return render(
            request,
            "coordinator/switch_teachers_list.html",
            {
                "page_type": "switch_teachers_list",
                "teachers": teachers_without_og,
            },
        )
    else:
        new_teacher_id = request.POST.get("new_teacher")
        new_teacher = TeacherInCharge.objects.filter(id=new_teacher_id).first()

        teacher_sessions = Teacher_Session.objects.filter(teacher=og_teacher)
        for teacher_session in teacher_sessions:
            session = teacher_session.session
            students = StudentsInfo.objects.filter(session=session, teacher=og_teacher)
            for student in students:
                student.teacher = new_teacher
                student_session = Student_Session.objects.filter(
                    session=session, student=student, teacher=og_teacher
                ).first()
                student_session.teacher = new_teacher
                student.save()
                student_session.save()
            teacher_session.teacher = new_teacher
            teacher_session.save()

        forms = FormDetails.objects.filter(teacher=og_teacher)
        for form in forms:
            form.teacher = new_teacher
            form.save()

        og_teacher_user.delete()
        request.session["my_messages"] = {"success": "Teacher switched successfully"}

        return redirect("accounts:coordinator_dashboard")


@login_required(login_url="accounts:loginlink")
@user_passes_test(is_coordinator, login_url="accounts:forbidden")
@password_change_required
def removeTeacher(request, teacher_id):
    teacher = TeacherInCharge.objects.filter(id=teacher_id).first()
    teacher_user = teacher.user
    organization = teacher.organization

    teacher_sessions = Teacher_Session.objects.filter(teacher=teacher)

    for teacher_session in teacher_sessions:
        session = teacher_session.session
        students = StudentsInfo.objects.filter(teacher=teacher, session=session)
        for student in students:
            student.teacher = None
            student.session = None
            student.save()
        teacher_session.delete()
        Student_Session.objects.filter(session=session, teacher=teacher).delete()

    teacher_user.delete()

    request.session["my_messages"] = {"success": "Teacher user deleted successfully"}
    return redirect("accounts:coordinator_dashboard")
